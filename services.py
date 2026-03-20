# services.py
import os
import re
import time
import io
import shutil
from pathlib import Path
from typing import Dict, List, Callable, Optional

from PIL import Image  # Para convertir GIF a PDF
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, JavascriptException, TimeoutException

from PyPDF2 import PdfMerger, PdfReader, PdfWriter

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from models import AppConfig, DayBatch, Order, OrderItem
import json
from config import (
    LIVERPOOL_ORDERS_URL,
    PENDING_TEXT,
    TIMEOUT,
)

# Ruta base de este archivo
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JS_DIR = os.path.join(BASE_DIR, "js")

with open(os.path.join(JS_DIR, "detail_items.js"), "r", encoding="utf-8") as f:
    DETAIL_ITEMS_JS = f.read()

with open(os.path.join(JS_DIR, "item_capture_rect.js"), "r", encoding="utf-8") as f:
    ITEM_CAPTURE_RECT_JS = f.read()


class LiverpoolService:
    """
    Capa de servicios:
    - Controla Selenium
    - Escanea /orders
    - Procesa detalles (dry-run)
    - Genera Excel + PDF
    - Acepta pedidos y descarga etiquetas (Fase 2)
    """

    def __init__(self, config: AppConfig, log_callback: Optional[Callable[[str], None]] = None):
        self.config = config
        self.log_callback = log_callback

    # ---------- logging ----------

    def log(self, msg: str):
        print(msg)
        if self.log_callback:
            self.log_callback(msg)

    # ---------- Selenium driver ----------

    def _init_driver(self) -> webdriver.Edge:
        """
        Inicializa Edge usando un perfil exclusivo para la automatización.
        - Usa un user-data-dir propio (C:\EdgeProfiles\LiverpoolAuto)
        - Reutiliza sesión después de la primera vez.
        """
        options = webdriver.EdgeOptions()

        # carpeta exclusiva para este bot
        options.add_argument(f"--user-data-dir={self.config.edge_user_data_dir}")
        # options.add_argument("--profile-directory=Default")
        options.add_argument("--start-maximized")

        # Intentar auto-detectar el driver con Selenium Manager (Selenium ≥ 4.6).
        # Si no funciona, caer al path fijo como fallback.
        FALLBACK_DRIVER = r"C:\WebDrivers\msedgedriver.exe"
        try:
            driver = webdriver.Edge(options=options)
            self.log("  [INFO] Edge driver detectado automáticamente por Selenium Manager.")
        except Exception:
            self.log(f"  [INFO] Usando driver fijo: {FALLBACK_DRIVER}")
            service = EdgeService(executable_path=FALLBACK_DRIVER)
            driver = webdriver.Edge(service=service, options=options)
        return driver

    def _set_page_size_250(self, driver: webdriver.Edge):
        """
        Intenta abrir el combo 'Resultados por página' y seleccionar 250.
        Usa el <div role="combobox" class="MuiTablePagination-select ...">.
        Si ya está en 250 o no lo encuentra, sólo registra un warning.
        """
        try:
            wait = WebDriverWait(driver, 10)

            combo = wait.until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//div[@role='combobox' and contains(@class,'MuiTablePagination-select')]",
                    )
                )
            )

            current = combo.text.strip()
            if current == "250":
                self.log("Resultados por página ya está en 250.")
                return

            # Abrir el desplegable
            wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//div[@role='combobox' and contains(@class,'MuiTablePagination-select')]",
                    )
                )
            ).click()

            # Seleccionar la opción 250
            opt_250 = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//li[@role='option' and normalize-space(.)='250']",
                    )
                )
            )
            opt_250.click()
            time.sleep(2)
            self.log("Se seleccionó '250' resultados por página.")
        except Exception as e:
            self.log(f"[WARN] No se pudo forzar 250 por página: {e}")

    # ---------- utilidades ----------

    def _parse_fecha_clave(self, fecha_texto: str) -> str:
        """
        Recibe algo como '17/11/2025 - 11:36 PM' y devuelve '2025-11-17'
        """
        if not fecha_texto:
            return ""
        parte_fecha = fecha_texto.split("-")[0].strip()
        m = re.search(r"(\d{1,2})\/(\d{1,2})\/(\d{4})", parte_fecha)
        if not m:
            return ""
        d, mth, y = m.groups()
        d = d.zfill(2)
        mth = mth.zfill(2)
        return f"{y}-{mth}-{d}"

    def _collect_orders_from_table(self, driver: webdriver.Edge) -> List[Order]:
        """
        Lee la tabla de la página actual de /orders
        y devuelve una lista de Order (sin items).
        """
        orders: List[Order] = []
        rows = driver.find_elements(By.XPATH, "//table//tbody//tr")

        for row in rows:
            try:
                link = row.find_element(By.XPATH, './/a[contains(@href,"/orders/detail/")]')
            except Exception:
                continue

            order_id = link.text.strip()
            url = link.get_attribute("href")

            tds = row.find_elements(By.TAG_NAME, "td")
            fecha_texto = ""
            for td in tds:
                txt = td.text.strip()
                if re.search(r"\d{1,2}\/\d{1,2}\/\d{4}", txt):
                    fecha_texto = txt
                    break

            fecha_clave = self._parse_fecha_clave(fecha_texto)

            estado = ""
            spans = row.find_elements(By.TAG_NAME, "span")
            for sp in spans:
                txt = sp.text.strip()
                if txt:
                    if "Pendiente" in txt or "Entregado" in txt or "Listo" in txt:
                        estado = txt
                        break

            orders.append(
                Order(
                    order_id=order_id,
                    url=url,
                    fecha_clave=fecha_clave,
                    fecha_texto=fecha_texto,
                    estado=estado,
                )
            )
        return orders

    # ---------- ESCANEO /orders ----------

    def scan_orders(self) -> Dict[str, DayBatch]:
        """
        Escanea /orders mientras haya al menos un 'Pendiente de aceptación'
        en la página actual. Se detiene cuando una página no tiene ninguno.
        Devuelve dict[fecha -> DayBatch] solo con pendientes.
        """
        days: Dict[str, DayBatch] = {}
        self.log("Iniciando escaneo de órdenes (Fase 1, sin aceptar)...")

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        try:
            driver.get(LIVERPOOL_ORDERS_URL)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))

            # Intentar fijar 250 por página
            self._set_page_size_250(driver)

            page_index = 1
            while True:
                self.log(f"Escaneando página {page_index}...")

                orders_page = self._collect_orders_from_table(driver)

                page_has_pending = False
                for o in orders_page:
                    if PENDING_TEXT in o.estado:
                        page_has_pending = True
                        if not o.fecha_clave:
                            continue
                        if o.fecha_clave not in days:
                            days[o.fecha_clave] = DayBatch(date=o.fecha_clave)
                        days[o.fecha_clave].orders.append(o)

                self.log(
                    f"Página {page_index}: {len(orders_page)} pedidos, "
                    f"{'con' if page_has_pending else 'sin'} pendientes de aceptación."
                )

                # Si ya no hay pendientes en esta página, paramos aquí
                if not page_has_pending:
                    self.log(
                        "Se encontró una página sin 'Pendiente de aceptación'. "
                        "Se detiene el paginado."
                    )
                    break

                # ---------- siguiente página ----------
                try:
                    wait_page = WebDriverWait(driver, 10)

                    next_btn = wait_page.until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//button[@aria-label="next page"]')
                        )
                    )

                    if not next_btn.is_enabled():
                        self.log("Botón 'siguiente' deshabilitado. Fin del paginado.")
                        break

                    driver.execute_script(
                        "arguments[0].scrollIntoView({block: 'center'});", next_btn
                    )
                    wait_page.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//button[@aria-label="next page"]')
                        )
                    )
                    driver.execute_script("arguments[0].click();", next_btn)

                    time.sleep(2)
                    page_index += 1

                except Exception as e:
                    self.log(
                        f"No se encontró botón 'siguiente'. Fin del paginado. Detalle: {e}"
                    )
                    break
                # -------------------------------------

            total_pend = sum(len(batch.orders) for batch in days.values())
            self.log(
                f"Escaneo completado. Fechas con pendientes: {len(days)}, "
                f"total pedidos pendientes: {total_pend}"
            )
            return days

        finally:
            driver.quit()

    def scan_orders_in_range(self, start_date_str: str, end_date_str: str) -> Dict[str, DayBatch]:
        """
        Escanea /orders buscando pedidos dentro del RANGO [start_date_str, end_date_str] (inclusive).
        Se detiene AL MOMENTO de encontrar una fecha MENOR a start_date_str (asumiendo orden descendente).
        """
        days: Dict[str, DayBatch] = {}
        self.log(f"Iniciando escaneo de rango: {start_date_str} al {end_date_str}")

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        try:
            driver.get(LIVERPOOL_ORDERS_URL)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            self._set_page_size_250(driver)

            page_index = 1
            stop_scan = False

            # Limite de seguridad alto por si acaso
            MAX_PAGES = 50 

            while page_index <= MAX_PAGES and not stop_scan:
                self.log(f"Escaneando página {page_index} (Rango {start_date_str} - {end_date_str})...")
                orders_page = self._collect_orders_from_table(driver)
                
                if not orders_page:
                    self.log("Página vacía. Fin del escaneo.")
                    break

                for o in orders_page:
                    if not o.fecha_clave:
                        continue
                    
                    # 1. Si fecha < start_date -> Ya nos pasamos, son muy viejos viajos. FIN.
                    if o.fecha_clave < start_date_str:
                        self.log(f"Fecha encontrada {o.fecha_clave} es menor al inicio del rango {start_date_str}. Deteniendo escaneo.")
                        stop_scan = True
                        break

                    # 2. Si fecha > end_date -> Es más nuevo de lo que queremos. Ignorar.
                    if o.fecha_clave > end_date_str:
                        continue

                    # 3. Si está en rango -> Guardar
                    if start_date_str <= o.fecha_clave <= end_date_str:
                        if o.fecha_clave not in days:
                            days[o.fecha_clave] = DayBatch(date=o.fecha_clave)
                        days[o.fecha_clave].orders.append(o)
                
                if stop_scan:
                    break

                # Siguiente página
                try:
                    wait_page = WebDriverWait(driver, 5)
                    next_btn = wait_page.until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//button[@aria-label="next page"]')
                        )
                    )
                    if not next_btn.is_enabled():
                        self.log("Botón Siguiente deshabilitado.")
                        break
                    
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
                    wait_page.until(EC.element_to_be_clickable((By.XPATH, '//button[@aria-label="next page"]')))
                    driver.execute_script("arguments[0].click();", next_btn)
                    time.sleep(2)
                    page_index += 1
                except Exception:
                    self.log("Fin de paginación o error al avanzar página.")
                    break

            total_orders = sum(len(b.orders) for b in days.values())
            self.log(f"Escaneo de rango completado. Encontrados {total_orders} pedidos en {len(days)} fechas dentro del rango.")
            return days

        finally:
            driver.quit()

    # ---------- DETALLES (dry-run) ----------

    def _get_order_id_from_detail(self, driver: webdriver.Edge) -> str:
        try:
            h4 = driver.find_element(
                By.XPATH, "//h4[contains(normalize-space(.),'Pedido n.')]"
            )
            txt = h4.text.strip()
            m = re.search(r"(\d{6,})", txt)
            return m.group(1) if m else ""
        except Exception:
            return ""

    def _is_order_pending(self, driver: webdriver.Edge) -> bool:
        """
        Verifica si en el detalle del pedido sigue apareciendo el estado
        'Pendiente de aceptación' usando el chip que muestras en el HTML.
        """
        try:
            driver.implicitly_wait(0)
            elem = driver.find_element(
                By.XPATH,
                "//span[contains(@class,'MuiChip-label') and "
                "contains(normalize-space(.),'Pendiente de aceptación')]",
            )
            return elem is not None
        except NoSuchElementException:
            return False
        finally:
            driver.implicitly_wait(0)

    def _get_items_from_detail(self, driver: webdriver.Edge) -> List[OrderItem]:
        """
        Obtiene los ítems del detalle ejecutando el JS de js/detail_items.js
        (es el mismo que usabas en consola).
        """
        raw_items = driver.execute_script(DETAIL_ITEMS_JS) or []
        self.log(f"  [DEBUG] JS devolvió {len(raw_items)} ítems.")

        items: List[OrderItem] = []

        for raw in raw_items:
            if not isinstance(raw, dict):
                continue

            title = (raw.get("title") or "").strip()
            if not title:
                continue

            qty_number = raw.get("qtyNumber")
            if qty_number is None:
                qty_number = 1

            try:
                qty_number = int(qty_number)
            except Exception:
                qty_number = 1

            item = OrderItem(
                title=title,
                qty=qty_number,
            )
            items.append(item)

        return items

    def _take_item_screenshots(self, driver: webdriver.Edge, order: Order, day_dir: Path):
        """
        Toma un screenshot recortado por ítem:
        - Incluye imagen, título y bloque 'Detalle del producto'.
        """
        if not order.items:
            return

        order_dir = day_dir / order.order_id
        order_dir.mkdir(parents=True, exist_ok=True)

        for item_idx, item in enumerate(order.items):
            img_path = order_dir / f"{order.order_id}_item{item_idx + 1}.png"
            ok = take_item_screenshot(driver, item_idx, str(img_path), logger=self.log)
            if not ok:
                self.log(
                    f"    [ERROR] Screenshot item {item_idx + 1} "
                    f"en pedido {order.order_id}"
                )
                continue

            # Una imagen por ítem; se reutiliza para las unidades (qty)
            item.screenshot_path = img_path

    def process_details_dry_run(self, days: Dict[str, DayBatch], selected_dates: List[str]):
        """
        Visita el detalle de cada pedido de las fechas seleccionadas.
        - NO hace clic en 'Aceptar'
        - NO entra a Documentos
        - NO descarga guías
        Solo lee info y toma screenshots por ítem, y genera archivos por día.
        """
        if not selected_dates:
            self.log("No hay fechas seleccionadas.")
            return

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        try:
            for date in selected_dates:
                batch = days.get(date)
                if not batch:
                    continue

                self.log(f"\n=== Procesando detalles (dry-run) para el día {date} ===")

                day_dir = self.config.base_dir / date
                day_dir.mkdir(parents=True, exist_ok=True)

                total_orders = len(batch.orders)

                for idx, order in enumerate(batch.orders, start=1):
                    self.log(f"[{date}] Pedido {order.order_id} ({idx}/{total_orders})")
                    try:
                        # Ir al detalle del pedido
                        driver.get(order.url)
                        wait.until(
                            EC.presence_of_element_located(
                                (
                                    By.XPATH,
                                    "//h4[contains(normalize-space(.),'Pedido n.')]",
                                )
                            )
                        )

                        # Validar que el ID de detalle coincide con el de la tabla
                        detail_order_id = self._get_order_id_from_detail(driver)
                        if detail_order_id and detail_order_id != order.order_id:
                            self.log(
                                f"  [WARN] ID detalle ({detail_order_id}) "
                                f"!= listado ({order.order_id}). "
                                "Marcando como error."
                            )
                            order.status = "error_id_mismatch"
                            continue

                        # --- Verificación robusta de 'Pendiente de aceptación' ---
                        try:
                            WebDriverWait(driver, TIMEOUT).until(
                                EC.presence_of_element_located(
                                    (
                                        By.XPATH,
                                        "//span[contains(@class,'MuiChip-label') "
                                        "and contains(normalize-space(.),"
                                        "'Pendiente de aceptación')]",
                                    )
                                )
                            )
                        except Exception:
                            self.log(
                                "  [INFO] El estado del pedido ya no es "
                                "'Pendiente de aceptación'. Se salta."
                            )
                            order.status = "skipped"
                            continue

                        chip_elems = driver.find_elements(
                            By.XPATH,
                            "//span[contains(@class,'MuiChip-label') "
                            "and contains(normalize-space(.),'Pendiente de aceptación')]",
                        )
                        if not chip_elems:
                            self.log(
                                "  [INFO] El estado del pedido ya no es "
                                "'Pendiente de aceptación'. Se salta."
                            )
                            order.status = "skipped"
                            continue

                        # --- Esperar a que carguen los ítems del pedido ---
                        try:
                            wait.until(
                                EC.presence_of_element_located(
                                    (
                                        By.CSS_SELECTOR,
                                        'div[class*="_OrderItem_item_name__"]',
                                    )
                                )
                            )
                        except Exception:
                            self.log(
                                "  [WARN] No se encontraron nodos de nombre de ítem "
                                "en el DOM. Se marca como error y se continúa."
                            )
                            order.status = "error_no_items_dom"
                            continue

                        # Obtener ítems del detalle
                        items = self._get_items_from_detail(driver)
                        if not items:
                            self.log("  [WARN] No se detectaron ítems en detalle.")
                        order.items = items

                        # Screenshots por ítem
                        self._take_item_screenshots(driver, order, day_dir)

                        order.status = "ok"

                    except Exception as e:
                        self.log(f"  [ERROR] Procesando pedido {order.order_id}: {e}")
                        order.status = "error"

                # Generar archivos (Excel + PDF) para este día
                self._generate_outputs_for_day(batch, day_dir)

        finally:
            driver.quit()

    # ---------- FASE 2: Aceptar pedidos + descargar etiquetas ----------

    def accept_and_download_labels(
        self,
        days: Dict[str, DayBatch],
        selected_dates: List[str],
    ):
        """
        Fase 2:
        - Para cada pedido de las fechas seleccionadas:
          * Abre el detalle del pedido
          * Si existe botón 'Aceptar', lo pulsa (y confirma si aplica)
          * Espera ~20s
          * Abre pestaña 'Documentos'
          * Descarga la 'Etiqueta de Envío'
          * Mueve el PDF a <base>/<fecha>/guias/<order_id>.pdf
        - Al final, une todas las guías del día en un solo PDF GUIAS_<fecha>.pdf
        """
        if not selected_dates:
            self.log("No hay fechas seleccionadas para Fase 2.")
            return

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        # Carpeta donde Edge deja las descargas
        download_dir: Path = getattr(
            self.config,
            "download_dir",
            Path.home() / "Downloads",
        )
        download_dir.mkdir(parents=True, exist_ok=True)

        try:
            for date in selected_dates:
                batch = days.get(date)
                if not batch:
                    continue

                self.log(f"\n=== Fase 2: Aceptar + descargar etiquetas para el día {date} ===")

                day_dir = self.config.base_dir / date
                day_dir.mkdir(parents=True, exist_ok=True)

                total_orders = len(batch.orders)

                for idx, order in enumerate(batch.orders, start=1):
                    self.log(f"[{date}] (F2) Pedido {order.order_id} ({idx}/{total_orders})")
                    try:
                        self._accept_and_download_for_order(
                            driver=driver,
                            wait=wait,
                            order=order,
                            day_dir=day_dir,
                            download_dir=download_dir,
                        )
                    except Exception as e:
                        self.log(f"  [ERROR F2] Pedido {order.order_id}: {e}")

                # Reintentar guías faltantes (1 vez)
                self._retry_missing_guides(batch, day_dir, driver)

                # Unir todas las guías de este día en un solo PDF (usando el orden de batch.orders)
                self._merge_labels_for_day(batch, day_dir, phase_label="F2")

        finally:
            self._cleanup_driver(driver)

    def reprocess_orders_execution(self, days: Dict[str, DayBatch], selected_dates: List[str]):
        """
        Reproceso de lista guardada:
        1. Recorre pedidos cargados.
        2. NO valida 'Pendiente de aceptación'.
        3. Toma screenshots de detalles.
        4. Descarga guías (si las hay).
        5. Genera Excel y PDF final.
        """
        if not selected_dates:
            self.log("No hay fechas para reprocesar.")
            return

        download_dir: Path = getattr(self.config, "download_dir", Path.home() / "Downloads")
        download_dir.mkdir(parents=True, exist_ok=True)

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        try:
            for date in selected_dates:
                batch = days.get(date)
                if not batch:
                    continue

                self.log(f"\n=== Reprocesando lista cargada para: {date} ===")
                day_dir = self.config.base_dir / date
                day_dir.mkdir(parents=True, exist_ok=True)

                total = len(batch.orders)
                for idx, order in enumerate(batch.orders, start=1):
                    self.log(f"[{date}] Reprocesando {order.order_id} ({idx}/{total})")

                    try:
                        # 1. Ir al detalle
                        driver.get(order.url)
                        wait.until(EC.presence_of_element_located((By.XPATH, "//h4[contains(normalize-space(.),'Pedido n.')]")))

                        # 2. Tomar screenshots (sin validar status)
                        try:
                            # Esperar carga items
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class*="_OrderItem_item_name__"]')))
                            order.items = self._get_items_from_detail(driver)
                            self._take_item_screenshots(driver, order, day_dir)
                            order.status = "ok"
                        except Exception as e:
                            self.log(f"  [WARN] Falló screenshot detalles: {e}")
                            order.status = "error_details"

                        # 3. Descargar guía (si existe)
                        try:
                            self._download_guide_only(driver, wait, order, day_dir, download_dir)
                        except Exception as e:
                            self.log(f"  [WARN] No se pudo descargar guía: {e}")

                    except Exception as e:
                        self.log(f"  [ERROR] Falló reproceso pedido {order.order_id}: {e}")

                # Reintentar guías faltantes
                self._retry_missing_guides(batch, day_dir, driver)

                # Generar reporte y unir PDFs
                self._generate_outputs_for_day(batch, day_dir)
                self._merge_labels_for_day(batch, day_dir, phase_label="REPROCESS")

        finally:
            self._cleanup_driver(driver)

    def _cleanup_driver(self, driver):
        try:
            driver.quit()
        except Exception:
            pass

    def save_orders_to_json(self, days: Dict[str, DayBatch], filepath: str = "orders_latest.json"):
        """
        Guarda el diccionario de días en un archivo JSON.
        """
        try:
            data = {date: batch.to_dict() for date, batch in days.items()}
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            self.log(f"Lista de pedidos guardada en: {filepath}")
        except Exception as e:
            self.log(f"[ERROR] No se pudo guardar JSON: {e}")

    def load_orders_from_json(self, filepath: str) -> Dict[str, DayBatch]:
        """
        Carga el diccionario de días desde un archivo JSON.
        """
        try:
            if not os.path.exists(filepath):
                self.log(f"[WARN] Archivo no encontrado: {filepath}")
                return {}

            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)

            days = {}
            for date, batch_data in data.items():
                days[date] = DayBatch.from_dict(batch_data)
            
            self.log(f"Lista cargada desde {filepath}. {len(days)} fechas encontradas.")
            return days
        except Exception as e:
            self.log(f"[ERROR] No se pudo cargar JSON: {e}")
            return {}

    def merge_labels_for_dates(
        self,
        days: Dict[str, DayBatch],
        selected_dates: List[str],
    ):
        """
        Fase 3 (botón 4):
        - Para cada fecha seleccionada:
          * Busca PDFs en <base_dir>/<fecha>/guias
          * Une las guías existentes en GUIAS_<fecha>.pdf
          * Genera GUIAS_FALTANTES_<fecha>.txt con los pedidos sin PDF
        """
        if not selected_dates:
            self.log("No hay fechas seleccionadas para unir guías.")
            return

        for date in selected_dates:
            batch = days.get(date)
            if not batch:
                self.log(f"[WARN F3] No hay batch para la fecha {date}.")
                continue

            day_dir = self.config.base_dir / date
            self.log(f"\n=== Fase 3: Unir guías para el día {date} ===")
            
            # Reintentar guías faltantes antes de unir
            # Nota: Esto abrirá navegador si hay faltantes
            self._retry_missing_guides(batch, day_dir, driver=None)
            
            self._merge_labels_for_day(batch, day_dir, phase_label="F3")

    def _retry_missing_guides(
        self,
        batch: DayBatch,
        day_dir: Path,
        driver: Optional[webdriver.Edge] = None,
    ):
        """
        Identifica pedidos sin PDF en <day_dir>/guias y reintenta descargarlos.
        Si 'driver' es None, se instancia uno nuevo (y se cierra al final).
        """
        guides_dir = day_dir / "guias"
        if not guides_dir.exists():
            guides_dir.mkdir(parents=True, exist_ok=True)

        # Identificar faltantes
        missing_orders = []
        for order in batch.orders:
            pdf_path = guides_dir / f"{order.order_id}.pdf"
            if not pdf_path.exists():
                missing_orders.append(order)

        if not missing_orders:
            self.log(f"  [INFO] No hay guías faltantes para {batch.date}. Todo completo.")
            return

        self.log(
            f"  [INFO] Se detectaron {len(missing_orders)} pedidos sin guía. "
            "Iniciando reintento..."
        )

        # Manejo del driver
        local_driver = False
        if driver is None:
            self.log("  [INFO] Iniciando navegador para reintentos...")
            driver = self._init_driver()
            local_driver = True
        
        wait = WebDriverWait(driver, TIMEOUT)
        download_dir: Path = getattr(
            self.config,
            "download_dir",
            Path.home() / "Downloads",
        )

        try:
            total = len(missing_orders)
            for idx, order in enumerate(missing_orders, start=1):
                self.log(f"  [RETRY] Procesando {order.order_id} ({idx}/{total})")
                try:
                    self._accept_and_download_for_order(
                        driver=driver,
                        wait=wait,
                        order=order,
                        day_dir=day_dir,
                        download_dir=download_dir,
                    )
                except Exception as e:
                    self.log(f"    [ERROR RETRY] Falló reintento de {order.order_id}: {e}")

        finally:
            if local_driver:
                driver.quit()
                self.log("  [INFO] Navegador de reintentos cerrado.")

    def process_old_orders_execution(self, days: Dict[str, DayBatch], selected_dates: List[str]):
        """
        Workflow de ejecución para 'Procesar Antiguos':
        Recorre cada fecha seleccionada y procesa sus pedidos:
           - Toma screenshots (detalle).
           - NO acepta pedidos.
           - Descarga guía (entra a documentos y baja).
        3. Genera excels, pdfs y merge de guías.
        """
        if not selected_dates:
            self.log("No hay fechas seleccionadas para procesar antiguos.")
            return

        download_dir: Path = getattr(self.config, "download_dir", Path.home() / "Downloads")
        download_dir.mkdir(parents=True, exist_ok=True)

        driver = self._init_driver()
        wait = WebDriverWait(driver, TIMEOUT)

        try:
            for date in selected_dates:
                batch = days.get(date)
                if not batch or not batch.orders:
                    self.log(f"No hay pedidos en memoria para la fecha {date}. Primero escanea.")
                    continue

                self.log(f"\n=== Procesando pedidos antiguos para {date} ===")
                day_dir = self.config.base_dir / date
                day_dir.mkdir(parents=True, exist_ok=True)

                total = len(batch.orders)
                for idx, order in enumerate(batch.orders, start=1):
                    self.log(f"[{date}] Procesando antiguo {order.order_id} ({idx}/{total})")
                    
                    try:
                        # A) Ir al detalle
                        driver.get(order.url)
                        wait.until(EC.presence_of_element_located(
                            (By.XPATH, "//h4[contains(normalize-space(.),'Pedido n.')]")
                        ))

                        # B) Tomar screenshots (lógica reutilizada de Dry-Run)
                        try:
                            # Esperar items
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class*="_OrderItem_item_name__"]')))
                            items_list = self._get_items_from_detail(driver)
                            order.items = items_list
                            # Screenshots
                            self._take_item_screenshots(driver, order, day_dir)
                            order.status = "ok" 
                        except Exception as e:
                            self.log(f"  [WARN] No se pudieron tomar screenshots: {e}")
                            order.status = "error_screenshots"

                        # C) Descargar guía (SIN aceptar)
                        self._download_guide_only(
                            driver=driver,
                            wait=wait,
                            order=order,
                            day_dir=day_dir,
                            download_dir=download_dir
                        )

                    except Exception as e:
                        self.log(f"  [ERROR] Falló proceso antiguo para {order.order_id}: {e}")

                # 3. Generar Outputs
                # Reintentar guías faltantes
                self._retry_missing_guides(batch, day_dir, driver)

                # Generar Excel y PDF visual
                self._generate_outputs_for_day(batch, day_dir)
                
                # Unir guías
                self._merge_labels_for_day(batch, day_dir, phase_label="OldOrders")

                self.log(f"=== Fin proceso antiguos para {date} ===")

        finally:
            driver.quit()

    def _download_guide_only(
        self,
        driver: webdriver.Edge,
        wait: WebDriverWait,
        order: Order,
        day_dir: Path,
        download_dir: Path
    ):
        """
        Entra a la pestaña documentos y descarga la guía.
        NO busca botones de aceptar.
        """
        # Abrir pestaña 'Documentos'
        try:
            documentos_tab = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//button[contains(@class,'MuiTab-root') "
                        "and contains(normalize-space(.),'Documentos')]",
                    )
                )
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", documentos_tab)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", documentos_tab)
            self.log("  [INFO Old] Pestaña 'Documentos' abierta.")
        except Exception as e:
            self.log(f"  [WARN Old] No se pudo abrir pestaña 'Documentos': {e}")
            return

        # Esperar tabla
        try:
            wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//table//tr[contains(@class,'MuiTableRow-root')]")
                )
            )
        except TimeoutException:
            self.log("  [WARN Old] No se encontraron filas en Documentos.")
            return

        # Descargar
        self._download_label_pdf(
            driver=driver,
            wait=wait,
            order=order,
            day_dir=day_dir,
            download_dir=download_dir,
        )

    def _merge_labels_for_day(self, batch: DayBatch, day_dir: Path, phase_label: str = "F3"):
        guides_dir = day_dir / "guias"
        if not guides_dir.exists():
            self.log(
                f"  [WARN {phase_label}] Carpeta de guías no existe para {batch.date}: {guides_dir}"
            )
            return

        merger = PdfMerger()
        added_any = False
        missing_orders: List[str] = []

        # Recorremos todos los pedidos del día
        for order in batch.orders:
            pdf_path = guides_dir / f"{order.order_id}.pdf"
            if pdf_path.exists():
                try:
                    merger.append(str(pdf_path))
                    added_any = True
                    self.log(
                        f"  [INFO {phase_label}] Añadiendo guía al PDF unificado: {pdf_path.name}"
                    )
                except Exception as e:
                    self.log(
                        f"  [WARN {phase_label}] No se pudo añadir {pdf_path.name} al unificado: {e}"
                    )
            else:
                missing_orders.append(order.order_id)

        if not added_any:
            self.log(
                f"  [WARN {phase_label}] No hay ninguna guía para unir en {guides_dir}."
            )
            merger.close()
            return

        merged_path = day_dir / f"GUIAS_{batch.date}.pdf"
        try:
            with open(merged_path, "wb") as f:
                merger.write(f)
            self.log(f"  [OK {phase_label}] PDF unificado de guías creado: {merged_path}")
        except Exception as e:
            self.log(f"  [WARN {phase_label}] No se pudo escribir el PDF unificado: {e}")
        finally:
            merger.close()

        # Excel con guías faltantes (en lugar de TXT)
        if missing_orders:
            missing_path = day_dir / f"GUIAS_FALTANTES_{batch.date}.xlsx"
            try:
                wb_missing = Workbook()
                ws_missing = wb_missing.active
                ws_missing.title = "Faltantes"
                ws_missing.append(["Pedido", "Link"])
                
                # Ajustar ancho de columnas
                ws_missing.column_dimensions["A"].width = 20
                ws_missing.column_dimensions["B"].width = 80

                for oid in missing_orders:
                    # Buscar el objeto order para obtener la URL
                    # (aunque es un poco ineficiente buscarlo de nuevo, es seguro)
                    found_order = next((o for o in batch.orders if o.order_id == oid), None)
                    url = found_order.url if found_order else ""
                    
                    ws_missing.append([oid, url])
                    
                    # Hacer el link clickable
                    if url:
                        cell = ws_missing.cell(row=ws_missing.max_row, column=2)
                        cell.hyperlink = url
                        cell.style = "Hyperlink"

                wb_missing.save(missing_path)
                self.log(
                    f"  [INFO {phase_label}] Archivo de guías faltantes generado: {missing_path}"
                )
            except Exception as e:
                self.log(
                    f"  [WARN {phase_label}] No se pudo escribir Excel de guías faltantes: {e}"
                )
        else:
            self.log(f"  [INFO {phase_label}] No hay pedidos sin PDF de guía en este día.")

    def _accept_and_download_for_order(
        self,
        driver: webdriver.Edge,
        wait: WebDriverWait,
        order: Order,
        day_dir: Path,
        download_dir: Path,
    ):
        """
        Abre el detalle, acepta el pedido si es posible y luego descarga la etiqueta.
        """
        # Ir al detalle
        driver.get(order.url)
        wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//h4[contains(normalize-space(.),'Pedido n.')]")
            )
        )

        # Intentar localizar el chip 'Pendiente de aceptación' (solo informativo)
        chip_xpath = (
            "//span[contains(@class,'MuiChip-label') "
            "and contains(normalize-space(.),'Pendiente de aceptación')]"
        )
        chip_present = False
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, chip_xpath)))
            chip_present = True
            self.log("  [INFO F2] Chip 'Pendiente de aceptación' visible.")
        except TimeoutException:
            self.log(
                "  [INFO F2] Chip 'Pendiente de aceptación' NO se vio con timeout; "
                "se intentará usar el botón 'Aceptar'."
            )

        # Buscar botón 'Aceptar' (si no está, asumimos que ya se aceptó)
        accept_btn = None
        try:
            accept_btn = wait.until(
                EC.presence_of_element_located((By.ID, "btn_order_accept"))
            )
        except TimeoutException:
            self.log(
                "  [INFO F2] No se encontró botón 'Aceptar' (btn_order_accept); "
                "se asume pedido ya aceptado y se pasa directo a 'Documentos'."
            )

        # Si no hay ni chip ni botón, asumimos que está en otro estado (ej. Pendiente de envío)
        # y continuamos para intentar descargar la guía.
        if not chip_present and accept_btn is None:
            self.log(
                "  [INFO F2] No hay chip 'Pendiente' ni botón 'Aceptar'. "
                "Se asume 'Pendiente de envío' y se continúa a Documentos."
            )
            # No hacemos return, dejamos que siga

        # Si hay botón, intentamos aceptar
        if accept_btn is not None:
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", accept_btn
                )
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", accept_btn)
                self.log("  [INFO F2] Click en botón 'Aceptar' (id=btn_order_accept).")

                # Si aparece un diálogo de confirmación, tratar de confirmarlo
                try:
                    confirm_btn = WebDriverWait(driver, 2).until(
                        EC.element_to_be_clickable(
                            (
                                By.XPATH,
                                "//button[.//span[contains(normalize-space(.),'Aceptar pedido')]]",
                            )
                        )
                    )
                    driver.execute_script("arguments[0].click();", confirm_btn)
                    self.log("  [INFO F2] Click en botón de confirmación 'Aceptar pedido'.")
                except TimeoutException:
                    # Puede que no haya confirmación y se acepte directo
                    self.log(
                        "  [INFO F2] No se encontró botón de confirmación 'Aceptar pedido'; "
                        "probablemente no hay diálogo."
                    )

                self.log("  [INFO F2] Esperando a que se procese el pedido...")
                time.sleep(10)

            except Exception as e:
                self.log(f"  [WARN F2] Error al intentar aceptar el pedido: {e}")

        # Abrir pestaña 'Documentos'
        try:
            documentos_tab = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//button[contains(@class,'MuiTab-root') "
                        "and contains(normalize-space(.),'Documentos')]",
                    )
                )
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", documentos_tab
            )
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", documentos_tab)
            self.log("  [INFO F2] Pestaña 'Documentos' abierta.")
        except Exception as e:
            self.log(f"  [WARN F2] No se pudo abrir pestaña 'Documentos': {e}")
            return

        # Esperar a que haya filas en la tabla de Documentos
        try:
            wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//table//tr[contains(@class,'MuiTableRow-root')]")
                )
            )
        except TimeoutException:
            self.log(
                "  [WARN F2] No se encontraron filas en la tabla de Documentos (timeout)."
            )
            return

        # Descargar la etiqueta de envío
        self._download_label_pdf(
            driver=driver,
            wait=wait,
            order=order,
            day_dir=day_dir,
            download_dir=download_dir,
        )

    def _download_label_pdf(
        self,
        driver: webdriver.Edge,
        wait: WebDriverWait,
        order: Order,
        day_dir: Path,
        download_dir: Path,
    ):
        """
        Hace clic en el botón de descarga dentro de la fila
        cuyo tipo de documento es 'Etiqueta de Envío' y mueve el PDF
        a <day_dir>/guias/<order_id>.pdf.

        IMPORTANTE:
        - Ya no intentamos cerrar pestañas nuevas (para evitar 'failed to close window')
        - Sólo observamos la carpeta de descargas para detectar el PDF nuevo.
        """
        guides_dir = day_dir / "guias"
        guides_dir.mkdir(parents=True, exist_ok=True)

        # PDFs o GIFs existentes antes del click
        before_pdfs = set(download_dir.glob("*.pdf"))
        before_gifs = set(download_dir.glob("*.gif"))

        try:
            # Fila cuyo tipo de documento es 'Etiqueta de Envío'
            row_xpath = (
                "//tr[contains(@class,'MuiTableRow-root') and "
                ".//div[contains(normalize-space(.),'Etiqueta de Envío')]]"
            )

            wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
            rows = driver.find_elements(By.XPATH, row_xpath)
            self.log(
                f"  [INFO F2] Filas con 'Etiqueta de Envío' encontradas en Documentos: {len(rows)}"
            )

            if not rows:
                self.log("  [WARN F2] No hay ninguna fila con 'Etiqueta de Envío'.")
                return

            row = rows[0]

            # --- DEBUG: HTML parcial de la fila ---
            try:
                row_html = row.get_attribute("innerHTML") or ""
                snippet = (row_html[:300] + "...") if len(row_html) > 300 else row_html
                self.log(f"  [DEBUG F2] HTML parcial de la fila 'Etiqueta de Envío': {snippet}")
            except Exception as e:
                self.log(f"  [DEBUG F2] No se pudo obtener innerHTML de la fila: {e}")

            # --- Buscar el botón de descarga por varios selectores ---
            btn = None
            tried_selectors = []

            # 1) SVG con data-testid tipo FileDownload
            sel1 = ".//svg[contains(@data-testid,'FileDownload')]/ancestor::button[1]"
            tried_selectors.append(sel1)
            try:
                elems1 = row.find_elements(By.XPATH, sel1)
                if elems1:
                    btn = elems1[0]
                    self.log(
                        f"  [INFO F2] Botón encontrado por selector 1 (svg[data-testid*='FileDownload']): {len(elems1)}"
                    )
            except Exception as e:
                self.log(f"  [DEBUG F2] Error usando selector 1: {e}")

            # 2) Cualquier botón tipo icon button dentro de la fila
            if btn is None:
                sel2 = ".//button[contains(@class,'MuiIconButton-root')]"
                tried_selectors.append(sel2)
                try:
                    elems2 = row.find_elements(By.XPATH, sel2)
                    if elems2:
                        btn = elems2[0]
                        self.log(
                            f"  [INFO F2] Botón encontrado por selector 2 (MuiIconButton-root): {len(elems2)}"
                        )
                except Exception as e:
                    self.log(f"  [DEBUG F2] Error usando selector 2: {e}")

            # 3) Como último recurso, cualquier <button> dentro de la fila
            if btn is None:
                sel3 = ".//button"
                tried_selectors.append(sel3)
                try:
                    elems3 = row.find_elements(By.XPATH, sel3)
                    if elems3:
                        btn = elems3[0]
                        self.log(
                            f"  [INFO F2] Botón encontrado por selector 3 (primer <button> de la fila): {len(elems3)}"
                        )

                        # DEBUG: loguear todos los botones de la fila
                        for i, b in enumerate(elems3, start=1):
                            try:
                                bh = b.get_attribute("outerHTML") or ""
                                snippet_btn = (
                                    bh[:200] + "..." if len(bh) > 200 else bh
                                )
                                self.log(
                                    f"    [DEBUG F2] Button #{i} en fila: {snippet_btn}"
                                )
                            except Exception:
                                pass
                except Exception as e:
                    self.log(f"  [DEBUG F2] Error usando selector 3: {e}")

            if btn is None:
                self.log(
                    f"  [WARN F2] No se encontró ningún botón de descarga en la fila "
                    f"'Etiqueta de Envío'. Selectores probados: {tried_selectors}"
                )
                return

            # --- Click al botón encontrado ---
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", btn
                )
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", btn)
                self.log(
                    "  [INFO F2] Click JS en botón de descarga dentro de la fila 'Etiqueta de Envío'."
                )
            except Exception as e:
                self.log(f"  [WARN F2] Error al hacer click en botón de descarga: {e}")
                return

        except Exception as e:
            self.log(
                f"  [WARN F2] Error localizando/clicando botón de descarga de etiqueta: {e}"
            )
            return

        # --- Esperar archivo nuevo (PDF o GIF) ---
        # Aumentamos timeout a 20s por si tarda en generar
        timeout = 20
        end_time = time.time() + timeout
        new_file = None

        while time.time() < end_time:
            current_pdfs = set(download_dir.glob("*.pdf"))
            current_gifs = set(download_dir.glob("*.gif"))

            new_pdfs = current_pdfs - before_pdfs
            new_gifs = current_gifs - before_gifs

            if new_pdfs:
                new_file = new_pdfs.pop()
                break
            if new_gifs:
                new_file = new_gifs.pop()
                break
            time.sleep(1)

        if not new_file:
            self.log("  [WARN F2] No se detectó archivo nuevo (PDF o GIF) tras el click.")
            return

        # --- Procesar el archivo descargado ---
        final_pdf_path = guides_dir / f"{order.order_id}.pdf"
        
        try:
            # Esperar a que termine de descargarse (tamaño estable)
            time.sleep(1.0) 
            
            # Si es GIF, convertir a PDF
            if new_file.suffix.lower() == ".gif":
                self.log(f"  [INFO F2] Se detectó guía en formato GIF: {new_file.name}")
                try:
                    with Image.open(new_file) as img:
                        # Rotar 90 grados a la derecha (clockwise) -> 270 grados counter-clockwise
                        # O usar transpose(Image.ROTATE_270)
                        rotated = img.transpose(Image.ROTATE_270)
                        
                        # Convertir a RGB para asegurar compatibilidad PDF
                        if rotated.mode != "RGB":
                            rotated = rotated.convert("RGB")
                        
                        rotated.save(final_pdf_path, "PDF", resolution=100.0)
                    
                    self.log(f"  [INFO F2] GIF convertido a PDF y guardado en: {final_pdf_path}")
                    
                    # Borrar el GIF original
                    try:
                        os.remove(new_file)
                    except Exception as e:
                        self.log(f"  [WARN F2] No se pudo borrar el GIF original: {e}")

                except Exception as e:
                    self.log(f"  [ERROR F2] Falló conversión de GIF a PDF: {e}")
                    return

            else:
                # Es PDF normal, moverlo
                shutil.move(str(new_file), str(final_pdf_path))
                self.log(f"  [INFO F2] PDF movido a: {final_pdf_path}")

        except Exception as e:
            self.log(f"  [ERROR F2] Error al procesar/mover el archivo descargado: {e}")




    # ---------- Generación de Excel / PDF ----------


    def _generate_outputs_for_day(self, batch: DayBatch, day_dir: Path):
        self.log(f"Generando archivos para el día {batch.date}...")

        ok_orders = [o for o in batch.orders if o.status == "ok"]
        if not ok_orders:
            self.log(f"  No hay pedidos 'ok' para {batch.date}. Se omiten archivos.")
            return

        # Excel principal
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        ws.column_dimensions["A"].width = 60
        row = 1

        for order in ok_orders:
            ws.cell(row=row, column=1, value=f"Pedido n. {order.order_id}")
            row += 1

            for item in order.items:
                for _ in range(item.qty):
                    if item.screenshot_path and item.screenshot_path.exists():
                        img = XLImage(str(item.screenshot_path))
                        ws.add_image(img, f"A{row}")
                    else:
                        ws.cell(
                            row=row,
                            column=1,
                            value=f"{item.title} (x{item.qty}) [SIN IMAGEN]",
                        )
                    row += 30

        excel_path = day_dir / f"PEDIDOS_{batch.date}.xlsx"
        wb.save(excel_path)
        self.log(f"  Excel principal: {excel_path}")

        # Excel múltiplos
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Control_multiples"
        ws2.append(["Pedido", "Cantidad total", "URL detalle"])

        for order in ok_orders:
            total_qty = sum(item.qty for item in order.items)
            if total_qty > 1:
                ws2.append([order.order_id, total_qty, order.url])

        multi_path = day_dir / f"MULTIPLES_{batch.date}.xlsx"
        wb2.save(multi_path)
        self.log(f"  Excel múltiplos: {multi_path}")

        # PDF A4 imprimible
        pdf_path = day_dir / f"PEDIDOS_{batch.date}_print.pdf"
        self._generate_print_pdf(ok_orders, pdf_path)
        self.log(f"  PDF imprimible: {pdf_path}")

    def _generate_print_pdf(self, ok_orders: List[Order], pdf_path: Path):
        """
        Genera PDF A4:
        - Encabezado 'Pedido n. XXXXX' una sola vez por pedido
        - Por unidad: título y debajo la imagen recortada (foto + detalle).
        Intentando meter ~2 pedidos por página.
        """
        page_width, page_height = A4
        margin_left = margin_right = 36
        margin_top = margin_bottom = 36

        max_width = page_width - margin_left - margin_right
        max_image_height = 260  # puntos aprox.

        c = canvas.Canvas(str(pdf_path), pagesize=A4)
        y = page_height - margin_top

        for order in ok_orders:
            # Expandimos por cantidad
            units: List[OrderItem] = []
            for item in order.items:
                for _ in range(item.qty):
                    units.append(item)

            if not units:
                continue

            order_header_drawn = False  # sólo una vez por pedido
            unit_index = 0

            for item in units:
                unit_index += 1

                # Caso sin imagen
                if not (item.screenshot_path and item.screenshot_path.exists()):
                    header_h = 18 if not order_header_drawn else 0
                    title_h = 14
                    block_height = header_h + title_h

                    if y - block_height < margin_bottom:
                        c.showPage()
                        y = page_height - margin_top

                    if not order_header_drawn:
                        c.setFont("Helvetica-Bold", 12)
                        c.drawString(margin_left, y, f"Pedido n. {order.order_id}")
                        y -= header_h
                        order_header_drawn = True

                    c.setFont("Helvetica", 10)
                    label = item.title
                    if len(units) > 1:
                        # Si son muchas unidades, mostramos que es la N de M
                        label_part = f" [SIN IMAGEN] (unidad {unit_index})"
                        c.drawString(margin_left, y, label + label_part)
                        
                        # --- MODIFICACION: Añadir leyenda roja ---
                        c.saveState()
                        c.setFillColorRGB(1, 0, 0) # Rojo
                        # La ponemos un poco a la derecha o debajo. 
                        # Para no romper layout, la ponemos a la derecha alineada (ej. x=350)
                        c.drawString(350, y, "SE REQUIERE TOMAR FOTO")
                        c.restoreState()
                        # -----------------------------------------
                    else:
                        label += " [SIN IMAGEN]"
                        c.drawString(margin_left, y, label)

                    y -= title_h + 10
                    continue

                # Hay screenshot recortado
                img = PILImage.open(str(item.screenshot_path))
                w, h = img.size

                if w > 0 and h > 0:
                    scale_w = max_width / float(w) if w > max_width else 1.0
                    scale_h = max_image_height / float(h) if h > max_image_height else 1.0
                    scale = min(scale_w, scale_h, 1.0)
                else:
                    scale = 1.0

                new_w = w * scale
                new_h = h * scale

                header_h = 18 if not order_header_drawn else 0
                title_h = 14
                block_height = header_h + title_h + new_h + 10

                # ¿Cabe el bloque en esta página?
                if y - block_height < margin_bottom:
                    c.showPage()
                    y = page_height - margin_top

                # Encabezado solo la primera vez de este pedido
                if not order_header_drawn:
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(margin_left, y, f"Pedido n. {order.order_id}")
                    y -= header_h
                    order_header_drawn = True

                # Título (opcionalmente marcamos unidad)
                c.setFont("Helvetica", 10)
                label = item.title
                if len(units) > 1:
                    label += f" (unidad {unit_index})"
                    
                    # --- MODIFICACION: Añadir leyenda roja ---
                    c.saveState()
                    c.setFillColorRGB(1, 0, 0) # Rojo
                    # La ponemos a la derecha alineada (ej. x=350)
                    c.drawString(350, y, "SE REQUIERE TOMAR FOTO")
                    c.restoreState()
                    # -----------------------------------------

                c.drawString(margin_left, y, label)
                y -= title_h

                # Imagen
                y -= new_h
                c.drawInlineImage(
                    img,
                    margin_left,
                    y,
                    width=new_w,
                    height=new_h,
                )
                y -= 10

            # Espacio extra entre pedidos
            y -= 20
            if y < margin_bottom:
                c.showPage()
                y = page_height - margin_top

        c.save()


# ---------- Función helper global para screenshots por ítem ----------

def take_item_screenshot(driver, item_index: int, out_path: str, logger=print) -> bool:
    """
    Toma un screenshot recortado del ítem:
    - Imagen
    - Título
    - Bloque 'Detalle del producto'
    Usa el rectángulo unificado devuelto por item_capture_rect.js.
    Si no se logra ajustar el rect completo al viewport, igualmente se
    usa el último rect_info válido para no perder la imagen.
    """
    try:
        # Asegurar que el ítem esté más o menos visible
        title_elems = driver.find_elements(
            By.CSS_SELECTOR, 'div[class*="_OrderItem_item_name__"]'
        )
        total_titles = len(title_elems)
        if item_index >= total_titles:
            logger(
                f"    [WARN] Índice {item_index} fuera de rango para títulos "
                f"({total_titles})"
            )
            return False

        # Intentar scrollear al contenedor raíz del ítem para que se vea completo
        try:
            title_el = title_elems[item_index]
            # Buscamos el ancestro 'MuiGrid-container' más cercano
            root_el = title_el.find_element(
                By.XPATH, "./ancestor::div[contains(@class, 'MuiGrid-container')][1]"
            )

            # 1. Intentar buscar la imagen y scrollear a ella para activar lazy loading
            try:
                img_el = root_el.find_element(By.CSS_SELECTOR, ".MuiCardMedia-root")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", img_el)
                time.sleep(0.5)  # Esperar carga de imagen
            except Exception:
                pass  # Si no hay imagen o falla, seguimos

            # 2. Scroll al top del elemento raíz para la captura
            driver.execute_script("arguments[0].scrollIntoView(true);", root_el)
            # Ajuste por header pegajoso (aprox 120px hacia arriba)
            driver.execute_script("window.scrollBy(0, -120);")

        except Exception as e:
            logger(f"    [DEBUG] Falló scroll al root, fallback a center title: {e}")
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});",
                title_elems[item_index],
            )

        # Aumentamos tiempo de espera para asegurar renderizado
        time.sleep(1.0)

        try:
            viewport_height = driver.execute_script(
                "return window.innerHeight || document.documentElement.clientHeight;"
            )
        except JavascriptException:
            viewport_height = 900  # fallback
    except JavascriptException as e:
        logger(f"    [ERROR] Al preparar scroll del ítem {item_index + 1}: {e}")
        return False

    rect_info = None

    # Hasta 3 intentos para que el rect completo quepa dentro del viewport
    for attempt in range(3):
        try:
            candidate = driver.execute_script(ITEM_CAPTURE_RECT_JS, item_index)
        except JavascriptException as e:
            logger(f"    [ERROR] JS rect ítem {item_index + 1}: {e}")
            return False

        logger(
            f"    [DEBUG] rect_info JS ítem {item_index + 1} (intento {attempt + 1}): {candidate}"
        )

        if not candidate:
            logger(f"    [WARN] JS devolvió null/undefined para ítem {item_index + 1}")
            return False

        if isinstance(candidate, dict) and candidate.get("ok") is False:
            logger(
                f"    [WARN] JS reportó error para ítem {item_index + 1}: {candidate}"
            )
            return False

        if not (isinstance(candidate, dict) and candidate.get("ok")):
            logger(
                f"    [WARN] Formato inesperado de rect_info para ítem {item_index + 1}: "
                f"type={type(candidate)}"
            )
            return False

        rect_info = candidate

        try:
            top = float(rect_info.get("top", 0))
            height = float(rect_info.get("height", 0))
        except Exception as e:
            logger(
                f"    [ERROR] No se pudieron leer coordenadas para ítem "
                f"{item_index + 1}: {e}"
            )
            return False

        bottom = top + height

        # ¿Ya cabe completo en el viewport?
        if top >= 0 and bottom <= viewport_height:
            break

        # Ajustar scroll y reintentar
        delta = 0
        if bottom > viewport_height:
            delta = bottom - viewport_height + 20  # bajar
        elif top < 0:
            delta = top - 20  # subir (valor negativo)

        logger(
            f"    [DEBUG] Ajustando scroll (delta={delta}) para ítem {item_index + 1}."
        )
        try:
            driver.execute_script("window.scrollBy(0, arguments[0]);", delta)
        except JavascriptException as e:
            logger(
                f"    [DEBUG] Error al hacer scroll de ventana para ítem "
                f"{item_index + 1}: {e}. Se usará el rect actual."
            )
            break

        time.sleep(0.4)
    else:
        logger(
            f"    [WARN] No se pudo ajustar el rect al viewport en 3 intentos "
            f"para ítem {item_index + 1}. Se usará el último rect disponible."
        )

    if rect_info is None:
        logger(
            f"    [WARN] No se obtuvo ningún rect_info válido para ítem "
            f"{item_index + 1}."
        )
        return False

    try:
        left = float(rect_info.get("left", 0))
        top = float(rect_info.get("top", 0))
        width = float(rect_info.get("width", 0))
        height = float(rect_info.get("height", 0))
    except Exception as e:
        logger(
            f"    [ERROR] No se pudieron leer coordenadas finales para ítem "
            f"{item_index + 1}: {e}"
        )
        return False

    if width <= 0 or height <= 0:
        logger(
            f"    [WARN] Rect final inválido para ítem {item_index + 1}: "
            f"left={left}, top={top}, width={width}, height={height}"
        )
        return False

    try:
        # Screenshot del viewport y recorte con Pillow
        png = driver.get_screenshot_as_png()
        img = PILImage.open(io.BytesIO(png))
        img_w, img_h = img.size

        l = max(0, int(left))
        t = max(0, int(top))
        r = min(img_w, int(left + width))
        b = min(img_h, int(top + height))

        if r <= l or b <= t:
            logger(
                f"    [WARN] Rect fuera de rango para ítem {item_index + 1}: "
                f"{(l, t, r, b)} (img: {img_w}x{img_h})"
            )
            return False

        cropped = img.crop((l, t, r, b))
        cropped.save(out_path)

        logger(
            f"    [INFO] Screenshot OK ítem {item_index + 1}: "
            f"{(l, t, r, b)} sobre imagen {img_w}x{img_h}"
        )
        return True

    except Exception as e:
        logger(f"    [ERROR] Al generar screenshot del ítem {item_index + 1}: {e}")
        return False
